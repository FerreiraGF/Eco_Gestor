import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# Caminho do arquivo de entrada e saída
arquivo = 'Base_Dados.xlsx'
saida = 'relatorio.xlsx'

# Verificar se o arquivo existe
if not os.path.exists(arquivo):
    print(f"Erro: O arquivo '{arquivo}' não foi encontrado.")
    exit()

# Carregar a planilha
try:
    df = pd.read_excel(arquivo)
except Exception as e:
    print(f"Erro ao carregar o arquivo: {e}")
    exit()

# Garantir que a coluna 'Data' esteja no formato datetime
try:
    df['Data'] = pd.to_datetime(df['Data'], format='%d/%m/%Y', errors='coerce')
    if df['Data'].isnull().any():
        print("Aviso: Algumas datas foram convertidas para 'NaT'. Verifique os dados de entrada.")
except KeyError:
    print("Erro: A coluna 'Data' não foi encontrada no arquivo.")
    exit()

# Extrair ano e mês 
df['Ano'] = df['Data'].dt.year
df['Mes'] = df['Data'].dt.month.astype(str).str.zfill(2)

# Função para sanitizar nomes de abas
def formatar_titulo_pl(titulo_pl):
    caracteres_invalidos = ['\\', '/', '*', '[', ']', ':', '?']
    for char in caracteres_invalidos:
        titulo_pl = titulo_pl.replace(char, '_')
    return titulo_pl

# Função para ajustar a largura das colunas
def ajustar_largura_colunas(planilha):
    for col in planilha.columns:
        comprimento_max = 0
        coluna = col[0].column_letter  # Obter o nome da coluna
        for cell in col:
            try:
                if len(str(cell.value)) > comprimento_max:
                    comprimento_max = len(cell.value)
            except:
                pass
        largura_ajustada = (comprimento_max + 2)
        planilha.column_dimensions[coluna].width = largura_ajustada

# Criar um escritor para Excel
with pd.ExcelWriter(saida, engine='openpyxl') as escritor:
    # Relatório por ano
    anos = sorted(df['Ano'].dropna().unique())
    
    # Criar relatório anual
    for ano in anos:
        df_ano = df[df['Ano'] == ano]
        relatorio_anual = (
            df_ano.groupby('Tipo de Resíduo')
            .agg(
                Peso_Total=('Peso (kg)', 'sum'),
                Quantidade=('Tipo de Resíduo', 'count')
            )
            .reset_index()
        )
        relatorio_anual['Frequencia'] = relatorio_anual['Quantidade'] / relatorio_anual['Quantidade'].sum()
        
        # Adicionar uma aba para o ano
        nome_aba = formatar_titulo_pl(f'{ano}')
        relatorio_anual.to_excel(escritor, sheet_name=nome_aba, index=False)

# Adicionar gráficos e consolidar relatórios mensais para os anos de 2022 e 2023
wb = load_workbook(saida)

for ano in [2022, 2023]:
    if str(ano) in wb.sheetnames:
        planilha = wb[str(ano)]
        
        # Gráfico de Porcentagem de Tipos de Resíduos
        df_tipo_residuo = df[df['Ano'] == ano].groupby('Tipo de Resíduo').agg(Quantidade=('Tipo de Resíduo', 'count')).reset_index()
        df_tipo_residuo['Porcentagem'] = df_tipo_residuo['Quantidade'] / df_tipo_residuo['Quantidade'].sum()
        
        grafico_pizza = PieChart()
        etiquetas = Reference(planilha, min_col=1, min_row=2, max_row=len(df_tipo_residuo) + 1)
        dados = Reference(planilha, min_col=3, min_row=1, max_row=len(df_tipo_residuo) + 1)
        grafico_pizza.add_data(dados, titles_from_data=True)
        grafico_pizza.set_categories(etiquetas)
        grafico_pizza.title = f'Porcentagem de Tipos de Resíduos - {ano}'
        grafico_pizza.style = 10  # Estilo do gráfico
        grafico_pizza.width = 20  # Largura do gráfico
        grafico_pizza.height = 10  # Altura do gráfico
        planilha.add_chart(grafico_pizza, 'E5')
        
        # Gráfico de Peso por Mês
        df_peso_mes = df[df['Ano'] == ano].groupby('Mes').agg(Peso_Total=('Peso (kg)', 'sum')).reset_index()
        for r in dataframe_to_rows(df_peso_mes, index=False, header=True):
            planilha.append(r)
        
        grafico_pizza_peso = PieChart()
        etiquetas_peso = Reference(planilha, min_col=1, min_row=planilha.max_row - len(df_peso_mes) + 1, max_row=planilha.max_row)
        dados_peso = Reference(planilha, min_col=2, min_row=planilha.max_row - len(df_peso_mes), max_row=planilha.max_row)
        grafico_pizza_peso.add_data(dados_peso, titles_from_data=True)
        grafico_pizza_peso.set_categories(etiquetas_peso)
        grafico_pizza_peso.title = f'Peso por Mês - {ano}'
        grafico_pizza_peso.style = 10  # Estilo do gráfico
        grafico_pizza_peso.width = 20  # Largura do gráfico
        grafico_pizza_peso.height = 10  # Altura do gráfico
        planilha.add_chart(grafico_pizza_peso, 'E20')
        
        # Consolidar relatórios mensais na planilha do ano
        df_ano_mes = df[df['Ano'] == ano].groupby(['Mes', 'Tipo de Resíduo']).agg(
            Peso_Total=('Peso (kg)', 'sum'),
            Quantidade=('Tipo de Resíduo', 'count')
        ).reset_index()
        
        linha_atual = planilha.max_row + 2 # Adicionar espaço entre as tabelas
        for mes in sorted(df[df['Ano'] == ano]['Mes'].unique()):
            df_mes = df_ano_mes[df_ano_mes['Mes'] == mes]
            planilha.append([f'Mês: {mes}'])
            for r in dataframe_to_rows(df_mes, index=False, header=True):
                planilha.append(r)
            linha_atual += len(df_mes) + 2  # Adicionar espaço entre as tabelas

        # Ajustar largura das colunas
        ajustar_largura_colunas(planilha)

        # Alinhar o texto das células
        for row in planilha.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

wb.save(saida)

print("Relatório gerado com sucesso!")