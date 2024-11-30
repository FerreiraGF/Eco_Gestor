import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from tkinter import *
from tkinter import messagebox
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from models.Usuario import Usuario

# Função para exibir os Termos de Uso
def exibir_termos():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    termos_path = os.path.join(current_dir, 'TermosUso.py')
    subprocess.Popen(['python', termos_path])

# Função para exibir a Política de Privacidade
def exibir_politica():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    politica_path = os.path.join(current_dir, 'politicaPrivacidade.py')
    subprocess.Popen(['python', politica_path])

# Função para cadastrar o usuário
def cadastrar():
    try:
        nome_empresa_valido = nome_empresa.get()
        cnpj_valido = Usuario.formatar_cnpj(cnpj.get())
        cep_valido = Usuario.formatar_cep(cep.get())
        email_valido = Usuario.formatar_email(email.get())
        telefone_valido = Usuario.formatar_telefone(telefone.get())
        senha_valida = Usuario.formatar_senha(senha.get())

        if senha.get() != confirmar_senha.get():
            raise ValueError("As senhas não coincidem!")

        if not termos_var.get():
            raise ValueError("Você deve concordar com os termos de uso e políticas de privacidade!")

        usuario = Usuario(nome_empresa_valido, cnpj_valido, cep_valido, email_valido, telefone_valido, senha_valida)

        # Salvar os Usuario cadastrados no Excel
        if os.path.exists('Usuarios_Cadastrados.xlsx'):
            wb = load_workbook('Usuarios_Cadastrados.xlsx')  # Carregar o workbook existente
            ws = wb.active  # Obter a planilha ativa
        else:
            wb = Workbook()  # Criar um novo workbook
            ws = wb.active  # Obter a planilha ativa
            # Adicionar cabeçalhos
            ws.append(["Nome da Empresa", "CNPJ", "CEP", "E-mail", "Telefone", "Senha"])
            # Formatar cabeçalhos em negrito
            for cell in ws[1]:
                cell.font = Font(bold=True)

        # Adicionar os dados do usuário
        ws.append([usuario.nome_empresa, usuario.cnpj, usuario.cep, usuario.email, usuario.telefone, usuario.senha])

        # Salvar o arquivo
        wb.save('Usuarios_Cadastrados.xlsx')
        messagebox.showinfo("Sucesso", "Cadastro realizado com sucesso!")
    except ValueError as e:
        messagebox.showerror("Erro", str(e))
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao salvar os dados: {e}")

# Criar a janela principal
janela = Tk()
janela.title("Cadastro")
janela.geometry("900x450")  # dimensionar a janela
frame = Frame(janela, padx=10, pady=10)
frame.pack(padx=10, pady=10)

Label(frame, text="Nome da Empresa:").grid(row=0, column=0, sticky="w")
nome_empresa = Entry(frame, width=50)
nome_empresa.grid(row=0, column=1, pady=5)

Label(frame, text="CNPJ:").grid(row=1, column=0, sticky="w")
cnpj = Entry(frame, width=50)
cnpj.grid(row=1, column=1, pady=5)

Label(frame, text="CEP:").grid(row=2, column=0, sticky="w")
cep = Entry(frame, width=50)
cep.grid(row=2, column=1, pady=5)

Label(frame, text="E-mail para Contato:").grid(row=3, column=0, sticky="w")
email = Entry(frame, width=50)
email.grid(row=3, column=1, pady=5)

Label(frame, text="Telefone:").grid(row=4, column=0, sticky="w")
telefone = Entry(frame, width=50)
telefone.grid(row=4, column=1, pady=5)

Label(frame, text="Senha:").grid(row=5, column=0, sticky="w")
senha = Entry(frame, show="*", width=50)
senha.grid(row=5, column=1, pady=5)

Label(frame, text="Confirmar Senha:").grid(row=6, column=0, sticky="w")
confirmar_senha = Entry(frame, show="*", width=50)
confirmar_senha.grid(row=6, column=1, pady=5)

termos_var = BooleanVar()
termos_check = Checkbutton(frame, text="Eu concordo com os", variable=termos_var)
termos_check.grid(row=7, column=0, sticky="w", pady=5)

# Frame para a frase interativa
termos_frame = Frame(frame, bg="#EAF7EC")
termos_frame.grid(row=7, column=1, columnspan=3, sticky="w")

# Adicionar palavras interativas para Termos de Uso e Política de Privacidade
termos_label = Label(termos_frame, text=" Termos de Uso ", fg="black", cursor="hand2", font=("Arial", 10, "underline"), bg="#EAF7EC")
termos_label.pack(side="left")
termos_label.bind("<Button-1>", lambda e: exibir_termos())

Label(termos_frame, text="e", font=("Arial", 10), bg="#EAF7EC").pack(side="left")

politica_label = Label(termos_frame, text=" Política de Privacidade", fg="black", cursor="hand2", font=("Arial", 10, "underline"), bg="#EAF7EC")
politica_label.pack(side="left")
politica_label.bind("<Button-1>", lambda e: exibir_politica())

Button(frame, text="Cadastrar", command=cadastrar).grid(row=8, column=0, columnspan=4, pady=10)

# Criar o rodapé
footer_frame = Frame(janela, bg="#2A5729", height=120)
footer_frame.pack(side="bottom", fill="x")

# Texto à esquerda do rodapé
left_text = Label(
    footer_frame,
    text="O Eco Gestor oferece uma solução inteligente para o\n"
         "descarte de resíduos eletrônicos, otimizando processos\n"
         "e garantindo a conformidade ambiental. Faça parte dessa\n"
         "luta ambiental conosco e ajude a mudar o planeta.",
    font=("Arial", 9),
    fg="white",
    bg="#2A5729",
    justify="left",
    padx=20,
    wraplength=350,
    anchor="w"
)
left_text.pack(side="left", anchor="w", padx=10)

# Texto à direita do rodapé
right_text = Label(
    footer_frame,
    text="Envie seu feedback para nós! Sua opinião nos ajuda a melhorar e transformar o planeta em um lugar melhor!\n eg_ouvidoria@ecogestor.com.br",
    font=("Arial", 9),
    fg="white",
    bg="#2A5729",
    justify="right",
    padx=20,
    wraplength=350,
    anchor="e"
)
right_text.pack(side="right", anchor="e", padx=10)

# Exibir a janela
janela.mainloop()